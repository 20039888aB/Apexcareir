import csv
import io
from decimal import Decimal

from django.db.models import DecimalField, F, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from apps.finance.models import Expense, Payroll
from apps.inventory.models import Product, StockReceipt
from apps.sales.models import Sale
from apps.common.services.company_branding import (
    branding_header_rows,
    draw_branded_pdf_header,
    get_company_branding,
    write_branded_xlsx_header,
)

REPORT_TYPES = {"sales", "inventory", "profit", "expenses", "performance"}


def _period_meta_rows(payload):
    rows = []
    if payload.get("period_label"):
        rows.append(["Period", payload["period_label"]])
    if payload.get("start_date"):
        rows.append(["Start Date", payload["start_date"]])
    if payload.get("end_date"):
        rows.append(["End Date", payload["end_date"]])
    if payload.get("generated_at"):
        rows.append(["Generated At", payload["generated_at"]])
    return rows


def _to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def build_report_payload(report_type, start_date, end_date, search):
    if report_type == "sales":
        return _sales_report(start_date, end_date, search)
    if report_type == "inventory":
        return _inventory_report(start_date, end_date, search)
    if report_type == "profit":
        return _profit_report(start_date, end_date, search)
    if report_type == "expenses":
        return _expense_report(start_date, end_date, search)
    if report_type == "performance":
        return _performance_report(start_date, end_date, search)
    raise ValueError("Invalid report type")


def _sales_report(start_date, end_date, search):
    queryset = Sale.objects.select_related("product", "salesperson").filter(date__gte=start_date, date__lte=end_date)
    if search:
        queryset = queryset.filter(
            Q(invoice_number__icontains=search)
            | Q(customer__icontains=search)
            | Q(product__name__icontains=search)
            | Q(product__sku__icontains=search)
        )

    rows = list(
        queryset.order_by("-date", "-created_at").values(
            "invoice_number",
            "customer",
            "product__name",
            "quantity",
            "price",
            "discount",
            "tax",
            "total",
            "profit",
            "salesperson__email",
            "date",
        )
    )
    total_sales = _to_float(
        queryset.aggregate(
            total=Coalesce(Sum("total"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )["total"]
    )
    total_profit = _to_float(
        queryset.aggregate(
            total=Coalesce(Sum("profit"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )["total"]
    )

    return {
        "title": "Sales Report",
        "filename": "sales-report",
        "summary": {
            "records": len(rows),
            "total_sales": total_sales,
            "total_profit": total_profit,
        },
        "columns": [
            {"key": "invoice_number", "label": "Invoice"},
            {"key": "customer", "label": "Customer"},
            {"key": "product__name", "label": "Product"},
            {"key": "quantity", "label": "Quantity"},
            {"key": "price", "label": "Price"},
            {"key": "discount", "label": "Discount"},
            {"key": "tax", "label": "Tax"},
            {"key": "total", "label": "Total"},
            {"key": "profit", "label": "Profit"},
            {"key": "salesperson__email", "label": "Salesperson"},
            {"key": "date", "label": "Date"},
        ],
        "results": rows,
    }


def _inventory_report(start_date, end_date, search):
    """Inventory activity for the selected period, driven by stock movements."""
    from apps.inventory.models import StockMovement

    movements = StockMovement.objects.select_related("product", "product__category", "product__supplier").filter(
        event_date__gte=start_date,
        event_date__lte=end_date,
    )
    if search:
        movements = movements.filter(
            Q(product__name__icontains=search)
            | Q(product__sku__icontains=search)
            | Q(reference_label__icontains=search)
            | Q(note__icontains=search)
        )

    rows = list(
        movements.order_by("-event_date", "-created_at").values(
            "event_date",
            "product__name",
            "product__sku",
            "product__category__name",
            "product__supplier__name",
            "movement_type",
            "quantity_change",
            "reference_label",
            "note",
            "product__current_stock",
            "product__minimum_stock",
        )
    )

    received_qty = sum(row["quantity_change"] for row in rows if row["quantity_change"] > 0)
    issued_qty = abs(sum(row["quantity_change"] for row in rows if row["quantity_change"] < 0))
    products_touched = {row["product__sku"] for row in rows}
    low_stock_count = Product.objects.filter(
        current_stock__lte=F("minimum_stock"),
        is_archived=False,
    ).exclude(status=Product.Status.DISCONTINUED).count()

    return {
        "title": "Inventory Report",
        "filename": "inventory-report",
        "summary": {
            "records": len(rows),
            "products_touched": len(products_touched),
            "quantity_received": received_qty,
            "quantity_issued": issued_qty,
            "low_stock_count": low_stock_count,
        },
        "columns": [
            {"key": "event_date", "label": "Date"},
            {"key": "product__name", "label": "Product"},
            {"key": "product__sku", "label": "SKU"},
            {"key": "product__category__name", "label": "Category"},
            {"key": "product__supplier__name", "label": "Supplier"},
            {"key": "movement_type", "label": "Movement"},
            {"key": "quantity_change", "label": "Qty Change"},
            {"key": "reference_label", "label": "Reference"},
            {"key": "note", "label": "Note"},
            {"key": "product__current_stock", "label": "Current Stock"},
            {"key": "product__minimum_stock", "label": "Minimum Stock"},
        ],
        "results": rows,
    }


def _profit_report(start_date, end_date, search):
    queryset = Sale.objects.select_related("product").filter(date__gte=start_date, date__lte=end_date)
    if search:
        queryset = queryset.filter(Q(invoice_number__icontains=search) | Q(customer__icontains=search))

    rows = list(
        queryset.order_by("-date").values(
            "invoice_number",
            "customer",
            "product__name",
            "quantity",
            "cost_price",
            "price",
            "total",
            "profit",
            "date",
        )
    )
    total_revenue = _to_float(
        queryset.aggregate(total=Coalesce(Sum("total"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )
    total_profit = _to_float(
        queryset.aggregate(total=Coalesce(Sum("profit"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )

    return {
        "title": "Profit Report",
        "filename": "profit-report",
        "summary": {
            "records": len(rows),
            "total_revenue": total_revenue,
            "total_profit": total_profit,
            "profit_margin_percent": round((total_profit / total_revenue * 100), 2) if total_revenue else 0,
        },
        "columns": [
            {"key": "invoice_number", "label": "Invoice"},
            {"key": "customer", "label": "Customer"},
            {"key": "product__name", "label": "Product"},
            {"key": "quantity", "label": "Quantity"},
            {"key": "cost_price", "label": "Cost Price"},
            {"key": "price", "label": "Selling Price"},
            {"key": "total", "label": "Total"},
            {"key": "profit", "label": "Profit"},
            {"key": "date", "label": "Date"},
        ],
        "results": rows,
    }


def _expense_report(start_date, end_date, search):
    queryset = Expense.objects.filter(date__gte=start_date, date__lte=end_date)
    if search:
        queryset = queryset.filter(Q(category__icontains=search) | Q(description__icontains=search))

    rows = list(
        queryset.order_by("-date").values(
            "category",
            "expense_type",
            "business_area",
            "amount",
            "description",
            "payment_method",
            "date",
        )
    )
    total_expenses = _to_float(
        queryset.aggregate(total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )

    return {
        "title": "Expense Report",
        "filename": "expense-report",
        "summary": {
            "records": len(rows),
            "total_expenses": total_expenses,
        },
        "columns": [
            {"key": "date", "label": "Date"},
            {"key": "expense_type", "label": "Expense Type"},
            {"key": "category", "label": "Category"},
            {"key": "business_area", "label": "Business Area"},
            {"key": "amount", "label": "Amount"},
            {"key": "payment_method", "label": "Payment Method"},
            {"key": "description", "label": "Description"},
        ],
        "results": rows,
    }


def _performance_report(start_date, end_date, search):
    sales_queryset = Sale.objects.filter(date__gte=start_date, date__lte=end_date)
    expense_queryset = Expense.objects.filter(date__gte=start_date, date__lte=end_date)
    payroll_queryset = Payroll.objects.filter(payment_date__gte=start_date, payment_date__lte=end_date)
    receipts_queryset = StockReceipt.objects.filter(date_received__gte=start_date, date_received__lte=end_date)

    total_revenue = _to_float(
        sales_queryset.aggregate(total=Coalesce(Sum("total"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )
    total_expenses = _to_float(
        expense_queryset.aggregate(total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )
    total_payroll = _to_float(
        payroll_queryset.aggregate(
            total=Coalesce(Sum("net_salary"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )["total"]
    )
    total_profit = _to_float(
        sales_queryset.aggregate(total=Coalesce(Sum("profit"), Value(0, output_field=DecimalField(max_digits=14, decimal_places=2))))[
            "total"
        ]
    )
    net_business_result = total_revenue - total_expenses - total_payroll

    results = [
        {
            "metric": "Total Revenue",
            "value": total_revenue,
        },
        {"metric": "Total Expenses", "value": total_expenses},
        {"metric": "Total Payroll", "value": total_payroll},
        {"metric": "Total Sales Profit", "value": total_profit},
        {"metric": "Net Business Result", "value": net_business_result},
        {"metric": "Sales Transactions", "value": sales_queryset.count()},
        {"metric": "Expense Records", "value": expense_queryset.count()},
        {"metric": "Payroll Records", "value": payroll_queryset.count()},
        {"metric": "Purchase Receipts", "value": receipts_queryset.count()},
    ]

    return {
        "title": "Business Performance Report",
        "filename": "business-performance-report",
        "summary": {
            "total_revenue": total_revenue,
            "total_expenses": total_expenses,
            "total_payroll": total_payroll,
            "net_business_result": net_business_result,
        },
        "columns": [
            {"key": "metric", "label": "Metric"},
            {"key": "value", "label": "Value"},
        ],
        "results": results,
    }


def export_as_csv(payload):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{payload["filename"]}.csv"'

    writer = csv.writer(response)
    for row in branding_header_rows():
        writer.writerow(row)
    writer.writerow([payload["title"]])
    for row in _period_meta_rows(payload):
        writer.writerow(row)
    writer.writerow([])
    writer.writerow(["Summary"])
    for key, value in payload["summary"].items():
        writer.writerow([key, value])
    writer.writerow([])
    writer.writerow([column["label"] for column in payload["columns"]])
    for row in payload["results"]:
        writer.writerow([row.get(column["key"], "") for column in payload["columns"]])
    return response


def export_as_xlsx(payload):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    next_row = write_branded_xlsx_header(worksheet, start_row=1)
    worksheet.cell(row=next_row, column=1, value=payload["title"])
    next_row += 1
    for row in _period_meta_rows(payload):
        worksheet.cell(row=next_row, column=1, value=row[0])
        worksheet.cell(row=next_row, column=2, value=row[1])
        next_row += 1
    next_row += 1
    worksheet.cell(row=next_row, column=1, value="Summary")
    next_row += 1
    for key, value in payload["summary"].items():
        worksheet.cell(row=next_row, column=1, value=key)
        worksheet.cell(row=next_row, column=2, value=value)
        next_row += 1
    next_row += 1
    for column_index, column in enumerate(payload["columns"], start=1):
        worksheet.cell(row=next_row, column=column_index, value=column["label"])
    next_row += 1
    for row in payload["results"]:
        for column_index, column in enumerate(payload["columns"], start=1):
            worksheet.cell(row=next_row, column=column_index, value=row.get(column["key"], ""))
        next_row += 1

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{payload["filename"]}.xlsx"'
    return response


def export_as_pdf(payload):
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    _, height = A4
    y = height - 40
    y = draw_branded_pdf_header(pdf, height, y)

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, payload["title"])
    y -= 20
    pdf.setFont("Helvetica", 9)
    for row in _period_meta_rows(payload):
        pdf.drawString(40, y, f"{row[0]}: {row[1]}")
        y -= 14
    y -= 10

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(40, y, "Summary")
    y -= 16
    pdf.setFont("Helvetica", 9)
    for key, value in payload["summary"].items():
        pdf.drawString(40, y, f"{key}: {value}")
        y -= 14
        if y < 80:
            pdf.showPage()
            y = height - 40
            pdf.setFont("Helvetica", 9)

    y -= 8
    pdf.setFont("Helvetica-Bold", 9)
    headers = " | ".join(column["label"] for column in payload["columns"])
    pdf.drawString(40, y, headers[:150])
    y -= 14
    pdf.setFont("Helvetica", 8)

    branding = get_company_branding()
    for row in payload["results"]:
        row_text = " | ".join(str(row.get(column["key"], "")) for column in payload["columns"])
        pdf.drawString(40, y, row_text[:150])
        y -= 12
        if y < 60:
            pdf.showPage()
            y = height - 40
            pdf.setFont("Helvetica", 8)
            y = draw_branded_pdf_header(pdf, height, y)

    pdf.setFont("Helvetica", 7)
    pdf.drawString(40, 30, f"{branding['company_name']} · Generated report")
    pdf.save()
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{payload["filename"]}.pdf"'
    return response
